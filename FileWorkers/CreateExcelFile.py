from datetime import date
from openpyxl import load_workbook
import openpyxl
import string
import os.path
import logging
from sys import exit


def create_excel_from_response(data_list, where_to_save):
    logging.info('Creating Excel file...')

    today = date.today()
    fmt_today = today.strftime("%Y.%m.%d")
    excel_filename = 'MSP_' + fmt_today + '.xlsx'
    # if file doesnt exist, create new. Менять заголовки если меняем поля местами, добавляем
    save_path = os.path.join(where_to_save, excel_filename)
    
    basedir = os.path.dirname(save_path)
    if not os.path.exists(basedir):
        os.makedirs(basedir)

    try:
        is_file_exist = os.path.isfile(save_path)
        if not is_file_exist:
            populate_headers(save_path)
    except Exception as e:
        logging.exception('Excel population error')
        exit(6)
    finally:
        wb = load_workbook(save_path)
        ws = wb.active

    # ws1 = wb.get_sheet_by_name("Sheet1")

    # Dict_headers = dict({'A1': 'Наименование / ФИО', 'B1': 'Тип субъекта', 'C1': 'Категория', 'D1': 'ОГРН',
    #                      'E1': 'ИНН', 'F1': 'Основной вид деятельности', 'G1': 'Регион', 'H1': 'Вновь созданный',
    #                      'I1': 'Дата включения в реестр', 'J1': 'Наличие лицензий',
    #                      'K1': 'Наличие заключенных договоров, контрактов',
    #                      'L1': 'Производство инновационной, высокотехнологичной продукции',
    #                      'M1': 'Участие в программах партнерства', 'N1': 'Является социальным предприятием',
    #                      'O1': 'Среднесписочная численность работников за предшествующий календарный год'})
    # file.append(Dict_headers)

    i = 2  # 2ая строка
    id_pp = 1
    letters = list(string.ascii_uppercase)

    try:
        for dat in data_list:
            for item in dat:
                name = item.get('name_ex', None)
                if name is None:
                    name = ''

                nptype = item.get('nptype', None)  # юр лицо или физ
                # TODO
                if nptype is None:
                    nptype = ''
                if nptype == "UL":
                    nptype = "Юридическое лицо"
                if nptype == "FL":
                    nptype = "Физическое лицо"
                if nptype == "IP":
                    nptype = "Индивидуальный предпрениматель"
                # else:
                #     nptype == ""
                #     logging.exception('NPTYPE is unknown')

                category = item.get('category', None)  # 1- Микропредприятие
                if category is None:
                    category = ''
                if category == 1:
                    category = "Микропредприятие"
                # TODO
                if category == 2:
                    category = "Малое предприятие"
                if category == 3:
                    category = "Среднее предприятие"
                if category == 4:
                    category = "Крупное предприятие"
                # else:
                #     logging.exception('Category is unknown')
                #     category = ""

                ogrn = item.get('ogrn', None)
                if ogrn is None:
                    ogrn = ''

                inn = item.get('inn', None)
                if inn is None:
                    inn = ''

                okved = item.get('okved1', None)  # Основной вид деятельности
                if okved is None:
                    okved = ''

                region = item.get('regioncode', None)
                if region is None:
                    region = ''

                isnew = item.get('isnew', None)  # Вновь созданный, если 0 - нет
                if isnew is None:
                    isnew = ''
                if isnew == 1:
                    isnew = "Вновь созданный"
                else:
                    isnew = "Нет"

                dtregistry = item.get('dtregistry', None)  # дата включения в реестр
                if dtregistry is None:
                    dtregistry = ''

                has_licenses = item.get('has_licenses', None)  # наличие лицензий
                if has_licenses is None:
                    has_licenses = ''
                if has_licenses == 0:
                    has_licenses = "Нет"
                else:
                    has_licenses = "Да"

                has_contracts = item.get('has_contracts', None)  # Наличие заключенных договоров, контрактов
                if has_contracts is None:
                    has_contracts = ''
                if has_contracts == 0:
                    has_contracts = "Нет"
                else:
                    has_contracts = "Да"

                is_hitech = item.get('is_hitech', None)  # Производство инновационной, высокотехнологичной продукции
                if is_hitech is None:
                    is_hitech = ''
                if is_hitech == 0:
                    is_hitech = "Нет"
                else:
                    is_hitech = "Да"

                is_partnership = item.get('is_partnership', None)  # Участие в программах партнерства
                if is_partnership is None:
                    is_partnership = ''
                if is_partnership == 0:
                    is_partnership = "Нет"
                else:
                    is_partnership = "Да"

                pr_soc = item.get('pr_soc', None)  # Является социальным предприятием
                if pr_soc is None:
                    pr_soc = ''
                if pr_soc == 0:
                    pr_soc = "Нет"
                else:
                    pr_soc = "Да"

                od2_sschr = item.get('od2_sschr', None)  # Среднесписочная численность работников за предшествующий
                if od2_sschr == 0 or od2_sschr is None:
                    od2_sschr = ''

                # календарный год

                # Dict = dict({'Наименование / ФИО': name, 'Тип субъекта': nptype, 'Категория': category, 'ОГРН': ogrn,
                #              'ИНН': inn, 'Основной вид деятельности': okved, 'Регион': region, 'Вновь созданный': isnew,
                #              'Дата включения в реестр': dtregistry, 'Наличие лицензий': has_licenses,
                #              'Наличие заключенных договоров, контрактов': has_contracts,
                #              'Производство инновационной, высокотехнологичной продукции': is_hitech,
                #              'Участие в программах партнерства': is_partnership, 'Является социальным предприятием': pr_soc,
                #              'Среднесписочная численность работников за предшествующий календарный год': od2_sschr})
                #
                # ws.append(Dict)

                # Alphabet Iterator
                for j in letters:
                    if j == 'A':
                        ws[j + str(i)] = id_pp

                    elif j == 'B':
                        ws[j + str(i)] = name
  
                    elif j == 'C':
                        ws[j + str(i)] = nptype

                    elif j == 'D':
                        ws[j + str(i)] = category

                    elif j == 'E':
                        ws[j + str(i)] = ogrn

                    elif j == 'F':
                        ws[j + str(i)] = inn

                    elif j == 'G':
                        ws[j + str(i)] = okved

                    elif j == 'H':
                        ws[j + str(i)] = region

                    elif j == 'I':
                        ws[j + str(i)] = isnew

                    elif j == 'J':
                        ws[j + str(i)] = dtregistry

                    elif j == 'K':
                        ws[j + str(i)] = has_licenses
  
                    elif j == 'L':
                        ws[j + str(i)] = has_contracts
     
                    elif j == 'M':
                        ws[j + str(i)] = is_hitech

                    elif j == 'N':
                        ws[j + str(i)] = is_partnership

                    elif j == 'O':
                        ws[j + str(i)] = pr_soc
     
                    elif j == 'P':
                        ws[j + str(i)] = od2_sschr



                i += 1
                id_pp += 1

        wb.save(save_path)

    except Exception as e:
        logging.exception("Excel file Exception")
        exit(7)
    logging.info('Excel file created')
    return save_path


def populate_headers(filename):
    i = 1  # 1ая строка
    letters = list(string.ascii_uppercase)
    wb = openpyxl.Workbook()
    ws = wb.active
    # Dict = dict({'Наименование / ФИО': name, 'Тип субъекта': nptype, 'Категория': category, 'ОГРН': ogrn,
    #              'ИНН': inn, 'Основной вид деятельности': okved, 'Регион': region, 'Вновь созданный': isnew,
    #              'Дата включения в реестр': dtregistry, 'Наличие лицензий': has_licenses,
    #              'Наличие заключенных договоров, контрактов': has_contracts,
    #              'Производство инновационной, высокотехнологичной продукции': is_hitech,
    #              'Участие в программах партнерства': is_partnership, 'Является социальным предприятием': pr_soc,
    #              'Среднесписочная численность работников за предшествующий календарный год': od2_sschr})
    #
    # Alphabet Iterator
    for j in letters:
        if j == 'A':
            ws[j + str(i)] = '№ п/п'
        elif j == 'B':
            ws[j + str(i)] = 'Наименование / ФИО'
        elif j == 'C':
            ws[j + str(i)] = 'Тип субъекта'
        elif j == 'D':
            ws[j + str(i)] = 'Категория'
        elif j == 'E':
            ws[j + str(i)] = 'ОГРН'
        elif j == 'F':
            ws[j + str(i)] = 'ИНН'
        elif j == 'G':
            ws[j + str(i)] = 'Основной вид деятельности'
        elif j == 'H':
            ws[j + str(i)] = 'Регион'
        elif j == 'I':
            ws[j + str(i)] = 'Вновь созданный'
        elif j == 'J':
            ws[j + str(i)] = 'Дата включения в реестр'
        elif j == 'K':
            ws[j + str(i)] = 'Наличие лицензий'
        elif j == 'L':
            ws[j + str(i)] = 'Наличие заключенных договоров, контрактов'
        elif j == 'M':
            ws[j + str(i)] = 'Производство инновационной, высокотехнологичной продукции'
        elif j == 'N':
            ws[j + str(i)] = 'Участие в программах партнерства'
        elif j == 'O':
            ws[j + str(i)] = 'Является социальным предприятием'
        elif j == 'P':
            ws[j + str(i)] = 'Среднесписочная численность работников за предшествующий календарный год'

    # Save all to Excel foreach line - add new
    wb.save(filename)

    i += 1
    return filename


# # Debug
# if __name__ == '__main__':
#  create_excel_from_response(resp,'')
